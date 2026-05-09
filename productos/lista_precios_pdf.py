"""PDF de lista de precios (membrete SIRONA + tabla). Compartido staff y portal vendedor."""

from __future__ import annotations

import math
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.contrib.staticfiles import finders
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from core.money_decimal import format_monto_ars
from core.pdf_membrete import platypus_membrete

from .models import ListaPrecioItem, ListaPrecios, Producto


def filas_lista_precios(lista: ListaPrecios) -> list[tuple[Producto, Decimal]]:
    # Farmacia/PDF: solo productos marcados en lista (igual que export y ficha).
    if lista.es_farmacia:
        qs = Producto.objects.filter(habilitado=True, en_lista_precios=True).order_by(
            "tipo", "descripcion", "codigo"
        )
        return [(p, p.precio_venta) for p in qs]
    items = (
        ListaPrecioItem.objects.filter(lista=lista, producto__habilitado=True)
        .select_related("producto")
        .order_by("producto__tipo", "producto__descripcion", "producto__codigo")
    )
    return [(i.producto, i.precio_venta) for i in items]


# PNG «para WhatsApp»: por debajo del umbral, una sola imagen; si es grande, el usuario elige categoría;
# dentro de cada categoría, las hojas reparten los renglones lo más parejo posible (sin una hoja casi vacía).
SPLIT_PNG_ROW_THRESHOLD = 200
PNG_MAX_ROWS_PER_IMAGE = 200

_TIPO_SORT_ORDER_PNG = {
    Producto.Tipo.MEDICAMENTOS: 0,
    Producto.Tipo.ACCESORIOS: 1,
    Producto.Tipo.OTROS: 2,
}

# Etiquetas de botón en la UI de export PNG (Medicamentos → Farmacia).
PNG_CATEGORY_BUTTON_LABELS = {
    Producto.Tipo.MEDICAMENTOS: "Farmacia",
    Producto.Tipo.ACCESORIOS: "Accesorios",
    Producto.Tipo.OTROS: "Otros",
}


def balanced_chunks(items: list[Any], max_per_chunk: int) -> list[list[Any]]:
    """
    Parte `items` en listas de tamaño similar, ninguna mayor que `max_per_chunk`.
    Ej.: 201 ítems y máx 200 → [101, 100], no [200, 1].
    """
    n = len(items)
    if n == 0:
        return []
    if n <= max_per_chunk:
        return [items]
    k = math.ceil(n / max_per_chunk)
    base = n // k
    rem = n % k
    out: list[list[Any]] = []
    idx = 0
    for i in range(k):
        sz = base + (1 if i < rem else 0)
        out.append(items[idx : idx + sz])
        idx += sz
    return out


def _payload_producto_png(p: Producto, precio: Decimal) -> dict[str, Any]:
    return {
        "codigo": p.codigo,
        "tipo": p.get_tipo_display(),
        "descripcion": p.descripcion,
        "precio": format_monto_ars(precio),
        "stock": int(p.stock or 0),
    }


def build_png_export_payload(filas: list[tuple[Producto, Decimal]]) -> dict[str, Any]:
    """
    Estructura para exportar PNG: modo único (lista corta) o por categoría con hojas balanceadas.
    """
    n = len(filas)
    if n <= SPLIT_PNG_ROW_THRESHOLD:
        return {
            "modo": "unico",
            "partes": [
                {
                    "titulo_suffix": "",
                    "filename_suffix": "",
                    "productos": [_payload_producto_png(p, pr) for p, pr in filas],
                    "hoja_num": 1,
                    "hojas_total": 1,
                }
            ],
            "categorias": [],
        }

    grouped: dict[str, list[tuple[Producto, Decimal]]] = {}
    for p, pr in filas:
        grouped.setdefault(p.tipo, []).append((p, pr))

    keys_sorted = sorted(grouped.keys(), key=lambda k: _TIPO_SORT_ORDER_PNG.get(k, 99))

    categorias: list[dict[str, Any]] = []
    for tipo_key in keys_sorted:
        rows = grouped[tipo_key]
        tipo_enum = Producto.Tipo(tipo_key)
        label_btn = PNG_CATEGORY_BUTTON_LABELS.get(tipo_enum, tipo_enum.label)
        slug_base = slugify(label_btn) or slugify(tipo_key) or "categoria"

        chunks = balanced_chunks(rows, PNG_MAX_ROWS_PER_IMAGE)
        partes_cat: list[dict[str, Any]] = []
        nh = len(chunks)
        for ci, chunk in enumerate(chunks):
            if nh == 1:
                titulo_suffix = label_btn
                fname = slug_base
            else:
                titulo_suffix = f"{label_btn} · Hoja {ci + 1}/{nh}"
                fname = f"{slug_base}-hoja-{ci + 1}"

            partes_cat.append(
                {
                    "titulo_suffix": titulo_suffix,
                    "filename_suffix": fname,
                    "productos": [_payload_producto_png(p, pr) for p, pr in chunk],
                    "hoja_num": ci + 1,
                    "hojas_total": nh,
                }
            )

        categorias.append(
            {
                "tipo_key": tipo_key,
                "button_label": label_btn,
                "productos_total": len(rows),
                "partes": partes_cat,
            }
        )

    return {
        "modo": "por_categoria",
        "partes": [],
        "categorias": categorias,
    }


def partes_lista_precios_png(filas: list[tuple[Producto, Decimal]]) -> list[dict[str, Any]]:
    """Compatibilidad: devuelve solo la lista de partes (modo único o vacío si es por categoría)."""
    payload = build_png_export_payload(filas)
    return list(payload.get("partes") or [])


def _truncate_text_to_width(
    value: object,
    max_width: float,
    *,
    font_name: str = "Helvetica",
    font_size: int = 9,
) -> str:
    text = str(value or "")
    if max_width <= 0:
        return ""
    if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
        return text

    suffix = "..."
    suffix_width = pdfmetrics.stringWidth(suffix, font_name, font_size)
    if suffix_width > max_width:
        return ""

    lo = 0
    hi = len(text)
    while lo < hi:
        mid = (lo + hi + 1) // 2
        candidate = text[:mid] + suffix
        if pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
            lo = mid
        else:
            hi = mid - 1
    return (text[:lo] + suffix) if lo > 0 else suffix


def _safe_filename(value: str, max_len: int = 60) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in value)[:max_len]


def lista_precios_pdf_file_response(*, lista: ListaPrecios) -> FileResponse:
    titulo = f"Lista de precios — {lista.nombre}"
    filas = filas_lista_precios(lista)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=18 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = platypus_membrete(titulo, doc.width, styles)

    tw = doc.width
    col_w = [tw * 0.15, tw * 0.18, tw * 0.45, tw * 0.22]
    cell_inner_w = [max(0, w - 12) for w in col_w]

    headers = ["Código", "Tipo", "Descripción", "Precio"]
    data = [headers]
    for p, precio in filas:
        data.append(
            [
                _truncate_text_to_width(p.codigo, cell_inner_w[0]),
                _truncate_text_to_width(p.get_tipo_display(), cell_inner_w[1]),
                _truncate_text_to_width(p.descripcion, cell_inner_w[2]),
                _truncate_text_to_width(format_monto_ars(precio), cell_inner_w[3]),
            ]
        )

    if len(data) == 1:
        data.append(["—", "—", "—", "—"])

    t = Table(data, colWidths=col_w, repeatRows=1)
    # Minimalista (sin "dashboard"): solo tabla limpia con encabezado sutil.
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LINEBELOW", (0, 0), (-1, 0), 0.75, colors.HexColor("#CBD5E1")),
                ("LINEABOVE", (0, 0), (-1, 0), 0.75, colors.HexColor("#CBD5E1")),
                ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(t)
    doc.build(story)
    buffer.seek(0)

    fecha = timezone.localtime().strftime("%d-%m-%Y")
    safe = _safe_filename(titulo)
    filename = f"{safe}_{fecha}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename, content_type="application/pdf")


def lista_precios_xlsx_response(*, lista: ListaPrecios) -> HttpResponse:
    titulo = f"Lista de precios — {lista.nombre}"
    filas = filas_lista_precios(lista)
    emitido = timezone.localtime()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de precios"
    ws.sheet_view.showGridLines = False

    sirona_teal = "0EA5A5"
    sirona_blue = "2563EB"
    header_fill = PatternFill("solid", fgColor=sirona_teal)
    title_fill = PatternFill("solid", fgColor=sirona_blue)
    soft_fill = PatternFill("solid", fgColor="EFF6FF")
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    border_color = "CBD5E1"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions["D"].width = 18

    for row in range(1, 5):
        ws.row_dimensions[row].height = 24
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = title_fill
            cell.font = Font(color="FFFFFF")

    ws.merge_cells("A1:D4")
    title_cell = ws["A1"]
    title_cell.value = titulo
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    logo_path = finders.find("img/sirona-logo.png")
    if logo_path:
        try:
            logo = XlsxImage(logo_path)
            original_width = logo.width
            original_height = logo.height
            logo.height = 46
            logo.width = int(original_width * (46 / original_height)) if original_height else original_width
            ws.add_image(logo, "A1")
        except Exception:
            pass

    ws.merge_cells("A5:D5")
    ws["A5"] = f"Emitido: {emitido.strftime('%d/%m/%Y %H:%M')} · Activa"
    ws["A5"].fill = soft_fill
    ws["A5"].font = Font(color="64748B", italic=True)
    ws["A5"].alignment = Alignment(horizontal="center")

    headers = ["Código", "Tipo", "Descripción", "Precio"]
    header_row = 7
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="right" if idx == 4 else "left")
        cell.border = thin_border

    row = header_row + 1
    if filas:
        for p, precio in filas:
            values = [p.codigo, p.get_tipo_display(), p.descripcion, Decimal(precio or 0)]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="right" if col == 4 else "left",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True,
                )
                if row % 2 == 1:
                    cell.fill = zebra_fill
                if col == 4:
                    cell.number_format = '"$" #,##0.00'
                    cell.font = Font(bold=True, color="0F172A")
            row += 1
    else:
        ws.cell(row=row, column=1, value="—")
        ws.cell(row=row, column=2, value="—")
        ws.cell(row=row, column=3, value="Sin productos")
        ws.cell(row=row, column=4, value="—")
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A{header_row}:D{max(header_row + 1, row - 1)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    for col in range(1, 5):
        ws.cell(row=header_row, column=col).border = thin_border
        ws.column_dimensions[get_column_letter(col)].bestFit = True

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = emitido.strftime("%d-%m-%Y")
    safe = _safe_filename(titulo)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{safe}_{fecha}.xlsx"'
    return resp
