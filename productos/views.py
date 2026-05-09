import math
import os
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models.deletion import ProtectedError
from django.db.models import (
    Avg,
    Case,
    CharField,
    Count,
    DecimalField,
    ExpressionWrapper,
    F,
    Q,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.http import FileResponse, HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from openpyxl import load_workbook
from urllib.parse import parse_qsl, urlencode

from core.export_utils import parse_export, pdf_response, xlsx_response
from core.authz import staff_required
from core.money_decimal import format_monto_ars, q2, redondear_precio_mostrador_ars
from core.pdf_membrete import platypus_membrete
from personas.models import Proveedor
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from .forms import ProductoForm
from .listas_precios_views import (
    producto_listas_extra_context,
    producto_listas_ids_post,
    producto_tiene_lista_precio_en_post,
    sync_producto_listas_extras_from_post,
)
from .models import ListaPrecioItem, ListaPrecios, Producto


def _productos_picker_data():
    """Lista compacta (código + descripción) para buscadores con menú desplegable (máx. 3000)."""
    return list(
        Producto.objects.order_by("descripcion", "codigo").values("codigo", "descripcion")[:3000]
    )


# Claves GET permitidas al volver a /productos/ (retorno_query o retorno_*).
_REDIR_PRODUCTOS_LIST_KEYS = frozenset(
    {"q", "tipo", "proveedor", "lista", "estado", "ingreso", "page", "ord", "dir"}
)

# Orden en listado: clave en URL -> campo en modelo
PRODUCTOS_SORT_FIELDS = {
    "codigo": "codigo",
    "descripcion": "descripcion",
    "tipo": "tipo",
    "costo": "costo",
    "stock": "stock",
    "pct": "porcentaje_ganancia",
    "precio": "precio_venta",
}


def _annotar_orden_tipo_alfa(qs):
    """Orden "alfabético" de rubro según el nombre en pantalla (no el código MED/AC/OT)."""
    return qs.annotate(
        _tipo_orden=Case(
            When(tipo=Producto.Tipo.MEDICAMENTOS, then=Value("Medicamentos")),
            When(tipo=Producto.Tipo.ACCESORIOS, then=Value("Accesorios")),
            When(tipo=Producto.Tipo.OTROS, then=Value("Otros")),
            default=Value(""),
            output_field=CharField(),
        )
    )


def _productos_ordenar_queryset(productos, request) -> tuple:
    """
    Sin ?ord= : rubro (nombre) → descripción A-Z → precio (e id).
    Con ?ord= & ?dir=asc|desc : una sola columna; mismo cabecero alterna asc/desc.
    """
    ord_key = (request.GET.get("ord") or "").strip()
    dir_raw = (request.GET.get("dir") or "").strip().lower()
    if dir_raw not in ("asc", "desc"):
        dir_raw = "asc"
    if ord_key in PRODUCTOS_SORT_FIELDS:
        f = PRODUCTOS_SORT_FIELDS[ord_key]
        prefix = "-" if dir_raw == "desc" else ""
        if f == "tipo":
            q2 = _annotar_orden_tipo_alfa(productos)
            return q2.order_by(f"{prefix}_tipo_orden", "id"), ord_key, dir_raw
        return productos.order_by(f"{prefix}{f}", "id"), ord_key, dir_raw
    qd = _annotar_orden_tipo_alfa(productos)
    return qd.order_by("_tipo_orden", "descripcion", "precio_venta", "id"), "", "asc"


def _productos_sort_links(request) -> dict[str, str]:
    """Querystring (sin page) para enlace de ordenar por cada columna."""
    out: dict[str, str] = {}
    for key in PRODUCTOS_SORT_FIELDS:
        q = request.GET.copy()
        cur_o = (request.GET.get("ord") or "").strip()
        cur_d = (request.GET.get("dir") or "asc").strip().lower()
        if cur_d not in ("asc", "desc"):
            cur_d = "asc"
        if cur_o == key:
            q["ord"] = key
            q["dir"] = "desc" if cur_d == "asc" else "asc"
        else:
            q["ord"] = key
            q["dir"] = "asc"
        q.pop("page", None)
        out[key] = q.urlencode()
    return out


def _productos_url_sin_orden_filtros(request) -> str:
    """Filtros actuales sin ord/dir/page (volver a orden predeterminado)."""
    q = request.GET.copy()
    q.pop("ord", None)
    q.pop("dir", None)
    q.pop("page", None)
    s = q.urlencode()
    return s


def _retorno_params_desde_post(request) -> dict:
    """
    Filtros para /productos/ leyendo retorno_query (misma query que la grilla) o, si no hay,
    los retorno_q / retorno_tipo / … (compatibilidad).
    """
    raw = (request.POST.get("retorno_query") or "").strip()
    if raw:
        out = {}
        for k, v in parse_qsl(raw, keep_blank_values=True):
            if k not in _REDIR_PRODUCTOS_LIST_KEYS or v is None or v == "":
                continue
            out[k] = v
        if out:
            return out
    params = {}
    for k in ("q", "tipo", "proveedor", "lista", "estado", "ingreso"):
        v = (request.POST.get(f"retorno_{k}") or "").strip()
        if v != "":
            params[k] = v
    pgn = (request.POST.get("retorno_page") or "").strip()
    if pgn.isdigit():
        params["page"] = pgn
    return params


def _redirect_productos_con_filtros(request):
    """Vuelve al listado con los mismos filtros que en la grilla (retorno_query o retorno_*)."""
    params = _retorno_params_desde_post(request)
    url = reverse("productos_list")
    if params:
        url += "?" + urlencode(params)
    return redirect(url)


def _filtrar_productos_queryset(request, *, use_post: bool = False):
    """Filtra por búsqueda, tipo y proveedor (productos con al menos una compra a ese proveedor)."""
    if use_post:
        q = (request.POST.get("filtro_q") or "").strip()
        tipo = (request.POST.get("filtro_tipo") or "").strip()
        proveedor = (request.POST.get("filtro_proveedor") or "").strip()
        lista = (request.POST.get("filtro_lista") or "").strip()
    else:
        q = (request.GET.get("q") or "").strip()
        tipo = (request.GET.get("tipo") or "").strip()
        proveedor = (request.GET.get("proveedor") or "").strip()
        lista = (request.GET.get("lista") or "").strip()

    estado = (request.GET.get("estado") or "").strip()
    ingreso = (request.GET.get("ingreso") or "").strip()

    productos = Producto.objects.all()
    if q:
        productos = productos.filter(Q(descripcion__icontains=q) | Q(codigo__icontains=q))
    if tipo:
        productos = productos.filter(tipo=tipo)
    if proveedor.isdigit():
        productos = productos.filter(compras_origen__proveedor_id=int(proveedor)).distinct()
    if lista.isdigit():
        lid = int(lista)
        lobj = ListaPrecios.objects.filter(pk=lid).first()
        if lobj:
            if lobj.es_farmacia:
                productos = productos.filter(en_lista_precios=True)
            else:
                productos = productos.filter(items_lista_precio__lista_id=lid).distinct()
    if estado == "1":
        productos = productos.filter(habilitado=True)
    elif estado == "0":
        productos = productos.filter(habilitado=False)
    if ingreso == "nuevos":
        desde = timezone.now() - timedelta(days=30)
        productos = productos.filter(creado_en__gte=desde)

    productos, sort_ord, sort_dir = _productos_ordenar_queryset(productos, request)

    return productos, {
        "q": q,
        "tipo": tipo,
        "proveedor": proveedor,
        "lista": lista,
        "estado": estado,
        "ingreso": ingreso,
        "sort_ord": sort_ord,
        "sort_dir": sort_dir,
    }


def _parse_pct_aumento(request) -> Decimal | None:
    raw = (request.POST.get("pct_aumento") or "").strip().replace(",", ".")
    if raw == "":
        return None
    try:
        d = Decimal(raw)
    except InvalidOperation:
        return None
    if d < 0 or d > Decimal("999.99"):
        return None
    return d


def _parse_precio_venta_input(raw: str) -> Decimal | None:
    s = (raw or "").strip().replace(",", ".")
    if s == "":
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    if d < 0:
        return None
    return q2(d)


def _redirect_productos_aumento_filtros(request):
    p = {
        "q": (request.POST.get("filtro_q") or "").strip(),
        "tipo": (request.POST.get("filtro_tipo") or "").strip(),
        "proveedor": (request.POST.get("filtro_proveedor") or "").strip(),
        "lista": (request.POST.get("filtro_lista") or "").strip(),
        "estado": (request.POST.get("filtro_estado") or "").strip(),
        "ingreso": (request.POST.get("filtro_ingreso") or "").strip(),
    }
    p = {k: v for k, v in p.items() if v != ""}
    qstr = urlencode(p) if p else ""
    url = reverse("productos_aumento")
    if qstr:
        url += "?" + qstr
    return redirect(url)


def _redirect_productos_lista_tras_aumento_guardado(request):
    """Vuelve al listado de productos con los mismos criterios que al armar el aumento."""
    p = {
        "q": (request.POST.get("filtro_q") or "").strip(),
        "tipo": (request.POST.get("filtro_tipo") or "").strip(),
        "proveedor": (request.POST.get("filtro_proveedor") or "").strip(),
        "lista": (request.POST.get("filtro_lista") or "").strip(),
        "estado": (request.POST.get("filtro_estado") or "").strip(),
        "ingreso": (request.POST.get("filtro_ingreso") or "").strip(),
    }
    p = {k: v for k, v in p.items() if v != ""}
    url = reverse("productos_list")
    if p:
        url += "?" + urlencode(p)
    return redirect(url)


def _celda_texto(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if math.isfinite(v) and abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return str(v).strip()
    return str(v).strip()


def _codigo_desde_celda(v) -> str:
    return _celda_texto(v)[:6]


def _parse_decimal_celda(v, *, default: Decimal) -> Decimal:
    if v is None:
        return default
    s = _celda_texto(v)
    if s == "":
        return default
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ValueError(f"valor numérico inválido ({s!r})") from exc


def _parse_opcional_decimal(v) -> Decimal | None:
    if v is None:
        return None
    s = _celda_texto(v)
    if s == "":
        return None
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ValueError(f"valor numérico inválido ({s!r})") from exc


def _sin_acentos(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _resolver_tipo_producto(tipo_raw: str) -> str | None:
    """
    Acepta mayúsculas/minúsculas, acentos y palabras parecidas
    (p. ej. medicamento/medicamentos, accesorio, otros).
    Devuelve Producto.Tipo.* o None.
    """
    s = _sin_acentos(tipo_raw.strip()).lower()
    s = " ".join(s.split())
    if not s:
        return None

    if s in ("med", "me"):
        return Producto.Tipo.MEDICAMENTOS
    if s == "ac":
        return Producto.Tipo.ACCESORIOS
    if s == "ot":
        return Producto.Tipo.OTROS

    if s.startswith("medic"):
        return Producto.Tipo.MEDICAMENTOS
    if s.startswith("acces"):
        return Producto.Tipo.ACCESORIOS

    if s in ("otros", "otro", "otr"):
        return Producto.Tipo.OTROS
    if s.startswith("otr") and not s.startswith("otra"):
        return Producto.Tipo.OTROS

    return None


# Importación Excel: índices por defecto (fila 1 = encabezados del modelo) y aliases de columnas
_IMPORT_COL_FIXED = {
    "codigo": 0,
    "descripcion": 1,
    "tipo": 2,
    "costo": 3,
    "porcentaje_ganancia": 4,
    "precio_venta": 5,
    "stock": 6,
    "fecha_vencimiento": 7,
}

_COLUMN_ALIAS_ORDER: list[tuple[str, frozenset[str]]] = [
    ("codigo", frozenset({"codigo", "code"})),
    ("descripcion", frozenset({"descripcion", "desc", "producto", "nombre"})),
    ("tipo", frozenset({"tipo", "type", "categoria", "rubro"})),
    ("costo", frozenset({"costo", "cost"})),
    (
        "porcentaje_ganancia",
        frozenset({"porcentaje_ganancia", "porcentaje", "ganancia", "margen", "%_gan", "%"}),
    ),
    ("precio_venta", frozenset({"precio_venta", "pvp", "precio"})),
    ("stock", frozenset({"stock", "cantidad", "unidades", "inv", "existencia"})),
    ("fecha_vencimiento", frozenset({"fecha_vencimiento", "vencimiento", "fecha_vto", "fecha"})),
]

_ALL_IMPORT_ALIASES: frozenset[str] = frozenset().union(*(a for _, a in _COLUMN_ALIAS_ORDER))


def _norm_encabezado_excel(cell) -> str:
    if cell is None:
        return ""
    s = str(cell).strip()
    if not s:
        return ""
    return _sin_acentos(s).lower().replace(" ", "_")


def _es_fila_encabezado_productos(cells: tuple) -> bool:
    """True si la fila parece títulos de columnas (no datos)."""
    hits = 0
    for c in cells:
        n = _norm_encabezado_excel(c)
        if n and n in _ALL_IMPORT_ALIASES:
            hits += 1
    return hits >= 2


def _construir_mapa_columnas_import(header_row: tuple) -> dict[str, int]:
    """Nombre lógico -> índice 0-based, según textos de la fila de encabezado."""
    colmap: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        n = _norm_encabezado_excel(cell)
        if not n:
            continue
        for key, aliases in _COLUMN_ALIAS_ORDER:
            if n in aliases and key not in colmap:
                colmap[key] = idx
                break
    return colmap


def _celda_import(row: tuple, key: str, colmap: dict[str, int] | None):
    # Si detectamos encabezados (colmap != None), usamos SOLO ese mapa.
    # Esto evita que, con encabezados parciales, se lean columnas equivocadas
    # (por ejemplo, que `stock` tome números de otra columna).
    if colmap is not None:
        if key not in colmap:
            return None
        i = colmap[key]
    else:
        i = _IMPORT_COL_FIXED[key]
    if len(row) <= i:
        return None
    return row[i]


def _parse_stock_importacion(v, fila_num: int) -> int:
    """
    Stock entero >= 0. No interpreta fechas ni números-serie de Excel como stock
    (evita cifras enormes si las columnas están desalineadas).
    """
    if v is None:
        return 0
    if isinstance(v, date):
        return 0
    if isinstance(v, bool):
        return 0
    if isinstance(v, (int, float)):
        fv = float(v)
        # Rango típico de serial de fecha (Excel): no usar como stock
        if 20000 < fv < 80000 and abs(fv - round(fv)) < 1e-9:
            return 0
    s = _celda_texto(v)
    if s == "":
        return 0
    try:
        n = int(_parse_decimal_celda(v, default=Decimal("0")))
    except ValueError as exc:
        raise ValueError(f"Fila {fila_num}: stock inválido ({s!r}).") from exc
    if n < 0:
        raise ValueError(f"Fila {fila_num}: el stock no puede ser negativo ({n}).")
    return n


IMPORT_EXCEL_CONFLICTS_KEY = "productos_import_excel_conflictos_v1"


def _tipo_label_producto(codigo_tipo: str) -> str:
    return dict(Producto.Tipo.choices).get(codigo_tipo, codigo_tipo)


def _excel_snapshot_for_session(
    defaults: dict,
    *,
    fecha_vencimiento: date | None,
) -> dict:
    """Valores serializables (JSON) para la sesión y para reaplicar si el usuario elige Excel."""
    snap: dict = {
        "descripcion": defaults["descripcion"],
        "tipo": defaults["tipo"],
        "tipo_label": _tipo_label_producto(defaults["tipo"]),
        "costo": str(defaults["costo"]),
        "stock": str(int(defaults["stock"])),
        "porcentaje_ganancia": str(defaults["porcentaje_ganancia"]),
        "fecha_vencimiento": fecha_vencimiento.isoformat() if fecha_vencimiento else "",
    }
    if defaults.get("precio_venta_editado"):
        snap["precio_venta"] = str(defaults["precio_venta"])
        snap["precio_automatico"] = False
    else:
        snap["precio_venta"] = ""
        snap["precio_automatico"] = True
    return snap


def _aplicar_snapshot_excel_a_producto(producto: Producto, snap: dict) -> None:
    producto.descripcion = (snap.get("descripcion") or "")[:255]
    producto.tipo = snap["tipo"]
    producto.costo = Decimal(snap["costo"])
    producto.stock = int(snap["stock"])
    producto.porcentaje_ganancia = Decimal(snap["porcentaje_ganancia"])
    fv = (snap.get("fecha_vencimiento") or "").strip()
    producto.fecha_vencimiento = date.fromisoformat(fv) if fv else None
    if snap.get("precio_automatico"):
        producto.precio_venta_editado = False
    else:
        producto.precio_venta = Decimal(snap["precio_venta"])
        producto.precio_venta_editado = True
    producto.save()


@login_required
def productos_list(request):
    productos, filtros_ctx = _filtrar_productos_queryset(request)
    q = filtros_ctx["q"]
    tipo = filtros_ctx["tipo"]
    proveedor = filtros_ctx["proveedor"]
    lista = filtros_ctx["lista"]
    estado = filtros_ctx["estado"]
    ingreso = filtros_ctx["ingreso"]
    sort_ord = filtros_ctx["sort_ord"]
    sort_dir = filtros_ctx["sort_dir"]

    exp = parse_export(request)
    if exp in ("xlsx", "pdf"):
        headers = [
            "Código",
            "Descripción",
            "Tipo",
            "Costo",
            "Stock",
            "% ganancia",
            "Precio venta",
            "Habilitado",
            "Lista precios",
            "Fecha vencimiento",
        ]
        rows = []
        for p in productos:
            rows.append(
                [
                    p.codigo,
                    p.descripcion,
                    p.get_tipo_display(),
                    str(p.costo),
                    p.stock,
                    str(p.porcentaje_ganancia),
                    str(p.precio_venta),
                    "Sí" if p.habilitado else "No",
                    "Sí" if p.en_lista_precios else "No",
                    p.fecha_vencimiento.strftime("%d/%m/%Y") if p.fecha_vencimiento else "",
                ]
            )
        base = "productos"
        if exp == "xlsx":
            return xlsx_response(base, [("Productos", headers, rows)])
        return pdf_response(base, "Listado de productos", [("Productos", headers, rows)])

    # KPIs del inventario (según filtros actuales).
    valor_total_expr = ExpressionWrapper(
        F("costo") * F("stock"),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )
    kpi = productos.aggregate(
        productos=Count("id"),
        activos=Count("id", filter=Q(habilitado=True)),
        stock_total=Coalesce(Sum("stock"), Value(0)),
        sin_stock=Count("id", filter=Q(stock__lte=0)),
        valor_total=Coalesce(Sum(valor_total_expr), Value(Decimal("0.00"))),
        margen_prom=Coalesce(Avg("porcentaje_ganancia"), Value(Decimal("0.00"))),
    )
    kpi["valor_total"] = q2(kpi.get("valor_total") or Decimal("0.00"))
    kpi["margen_prom"] = q2(kpi.get("margen_prom") or Decimal("0.00"))

    page = (request.GET.get("page") or "").strip()
    paginator = Paginator(productos, 100)
    page_obj = paginator.get_page(page or 1)
    productos_page = list(page_obj)
    qcopy = request.GET.copy()
    qcopy.pop("page", None)
    querystring = qcopy.urlencode()

    proveedores_filtro = Proveedor.objects.filter(habilitado=True).order_by("apellido", "nombre", "codigo")
    listas_precios_filtro = ListaPrecios.objects.all().order_by("-es_farmacia", "nombre")
    proveedor_sel_label = ""
    if proveedor.isdigit():
        pr = Proveedor.objects.filter(pk=int(proveedor)).values_list("apellido", "nombre").first()
        if pr:
            proveedor_sel_label = f"{pr[0]}, {pr[1]}".strip().strip(",")

    return render(
        request,
        "productos/list.html",
        {
            "productos": productos_page,
            "q": q,
            "tipo": tipo,
            "proveedor": proveedor,
            "lista": lista,
            "estado": estado,
            "ingreso": ingreso,
            "sort_ord": sort_ord,
            "sort_dir": sort_dir,
            "sort_links": _productos_sort_links(request),
            "querystring_no_sort": _productos_url_sin_orden_filtros(request),
            "tipos": Producto.Tipo.choices,
            "proveedores_filtro": proveedores_filtro,
            "listas_precios_filtro": listas_precios_filtro,
            "proveedor_sel_label": proveedor_sel_label,
            "page_obj": page_obj,
            "querystring": querystring,
            "kpi": kpi,
            "productos_picker": _productos_picker_data(),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def productos_aumento(request):
    proveedores_filtro = Proveedor.objects.filter(habilitado=True).order_by("apellido", "nombre", "codigo")

    if request.method == "POST":
        step = (request.POST.get("step") or "").strip()
        if step == "confirm":
            pct = _parse_pct_aumento(request)
            ids = [int(x) for x in request.POST.getlist("producto_id") if str(x).isdigit()]
            if pct is None:
                messages.error(request, "El porcentaje de aumento no es válido.")
                return _redirect_productos_aumento_filtros(request)
            if not ids:
                messages.error(request, "No se recibieron productos para actualizar.")
                return _redirect_productos_aumento_filtros(request)

            for sid in ids:
                if _parse_precio_venta_input(request.POST.get(f"precio_{sid}") or "") is None:
                    messages.error(
                        request,
                        f"Revisá el precio final del producto #{sid} (debe ser un importe válido).",
                    )
                    return _redirect_productos_aumento_filtros(request)

            factor = Decimal("1.0") + (pct / Decimal("100"))
            actualizados = 0
            try:
                with transaction.atomic():
                    for sid in ids:
                        precio = _parse_precio_venta_input(request.POST.get(f"precio_{sid}") or "")
                        p = Producto.objects.select_for_update().get(pk=sid)
                        p.costo = q2(p.costo * factor)
                        p.precio_venta = precio
                        p.precio_venta_editado = True
                        p.save()
                        actualizados += 1
            except Producto.DoesNotExist:
                messages.error(request, "Algún producto ya no existe.")
                return _redirect_productos_aumento_filtros(request)

            messages.success(
                request,
                f"Aumento del {pct}% aplicado sobre el costo en {actualizados} producto(s).",
            )
            return _redirect_productos_lista_tras_aumento_guardado(request)

        if step == "preview":
            ids = [int(x) for x in request.POST.getlist("sel") if str(x).isdigit()]
            pct = _parse_pct_aumento(request)
            if not ids:
                messages.error(request, "Seleccioná al menos un producto.")
                return _redirect_productos_aumento_filtros(request)
            if pct is None:
                messages.error(request, "Indicá un porcentaje de aumento válido (ej.: 10 para 10%).")
                return _redirect_productos_aumento_filtros(request)

            factor = Decimal("1.0") + (pct / Decimal("100"))
            productos_sel = (
                Producto.objects.filter(pk__in=ids).order_by("descripcion", "codigo")
            )
            rows = []
            for p in productos_sel:
                nuevo_costo = q2(p.costo * factor)
                sugerido = redondear_precio_mostrador_ars(
                    nuevo_costo
                    * (Decimal("1.0") + (p.porcentaje_ganancia / Decimal("100")))
                )
                rows.append(
                    {
                        "producto": p,
                        "costo_anterior": p.costo,
                        "nuevo_costo": nuevo_costo,
                        "precio_sugerido": sugerido,
                    }
                )

            fq = (request.POST.get("filtro_q") or "").strip()
            ft = (request.POST.get("filtro_tipo") or "").strip()
            fp = (request.POST.get("filtro_proveedor") or "").strip()
            fl = (request.POST.get("filtro_lista") or "").strip()
            fe = (request.POST.get("filtro_estado") or "").strip()
            fi = (request.POST.get("filtro_ingreso") or "").strip()
            back_q = {}
            if fq:
                back_q["q"] = fq
            if ft:
                back_q["tipo"] = ft
            if fp:
                back_q["proveedor"] = fp
            if fl:
                back_q["lista"] = fl
            if fe:
                back_q["estado"] = fe
            if fi:
                back_q["ingreso"] = fi
            aumento_back_url = reverse("productos_aumento")
            if back_q:
                aumento_back_url += "?" + urlencode(back_q)

            return render(
                request,
                "productos/aumento.html",
                {
                    "step": "confirm",
                    "pct": pct,
                    "rows": rows,
                    "filtro_q": fq,
                    "filtro_tipo": ft,
                    "filtro_proveedor": fp,
                    "filtro_lista": fl,
                    "filtro_estado": fe,
                    "filtro_ingreso": fi,
                    "aumento_back_url": aumento_back_url,
                    "proveedores_filtro": proveedores_filtro,
                },
            )

    productos, filtros_ctx = _filtrar_productos_queryset(request)
    listas_filtro = ListaPrecios.objects.all().order_by("-es_farmacia", "nombre")
    return render(
        request,
        "productos/aumento.html",
        {
            "step": "filter",
            "productos": productos,
            "q": filtros_ctx["q"],
            "tipo": filtros_ctx["tipo"],
            "proveedor": filtros_ctx["proveedor"],
            "lista": filtros_ctx["lista"],
            "estado": filtros_ctx["estado"],
            "ingreso": filtros_ctx["ingreso"],
            "tipos": Producto.Tipo.choices,
            "proveedores_filtro": proveedores_filtro,
            "listas_precios_filtro": listas_filtro,
            "productos_picker": _productos_picker_data(),
        },
    )


@login_required
@require_http_methods(["POST"])
def lista_precios_guardar(request):
    messages.info(
        request,
        "Las listas de precios se administran desde la pestaña Listas de precio (rubros y precios por lista).",
    )
    return redirect("productos_listas_precios")


@login_required
@require_http_methods(["POST"])
def lista_precios_aplicar(request):
    messages.info(
        request,
        "Para marcar productos en el PDF de Farmacia usá el interruptor «En lista (PDF)» en cada producto.",
    )
    return redirect("productos_listas_precios")


def _producto_listas_seleccionadas_post(request) -> set[int] | None:
    if request.method != "POST" or request.POST.get("listas_extra_present") != "1":
        return None
    return {int(x) for x in request.POST.getlist("listas_extra") if str(x).isdigit()}


def _producto_requiere_lista_precio(request, form) -> bool:
    if request.POST.get("listas_extra_present") != "1":
        return False
    if producto_tiene_lista_precio_en_post(request):
        return False
    msg = (
        "Asigná el producto a al menos una lista de precio. "
        "Si no lo hacés, no aparece en Farmacia ni en ningún rubro para vender o exportar."
    )
    form.add_error(None, msg)
    messages.error(request, msg)
    return True


def _normalizar_descripcion_producto(valor: str) -> str:
    return " ".join(_sin_acentos(valor or "").casefold().split())


def _producto_listas_ids_actuales(producto: Producto, farmacia_id: int | None) -> set[int]:
    ids = set(
        ListaPrecioItem.objects.filter(producto_id=producto.pk).values_list("lista_id", flat=True)
    )
    if farmacia_id and producto.en_lista_precios:
        ids.add(farmacia_id)
    return ids


def _producto_repite_en_misma_lista(request, form, producto: Producto | None = None) -> bool:
    if request.POST.get("listas_extra_present") != "1":
        return False

    descripcion = form.cleaned_data.get("descripcion") if hasattr(form, "cleaned_data") else ""
    descripcion_norm = _normalizar_descripcion_producto(str(descripcion or ""))
    if not descripcion_norm:
        return False

    listas_seleccionadas = producto_listas_ids_post(request)
    if not listas_seleccionadas:
        return False

    farmacia_id = (
        ListaPrecios.objects.filter(es_farmacia=True)
        .order_by("id")
        .values_list("pk", flat=True)
        .first()
    )
    listas_nombres = dict(ListaPrecios.objects.values_list("pk", "nombre"))
    candidatos = Producto.objects.all()
    if producto and producto.pk:
        candidatos = candidatos.exclude(pk=producto.pk)

    for candidato in candidatos:
        if _normalizar_descripcion_producto(candidato.descripcion) != descripcion_norm:
            continue
        repetidas = listas_seleccionadas & _producto_listas_ids_actuales(candidato, farmacia_id)
        if not repetidas:
            continue
        nombres = ", ".join(sorted(listas_nombres.get(pk, str(pk)) for pk in repetidas))
        msg = (
            f"Ya existe un producto con esa descripción ({candidato.codigo}) en: {nombres}. "
            "Solo se permite repetirlo si pertenece a listas de precio distintas."
        )
        form.add_error("descripcion", msg)
        messages.error(request, msg)
        return True

    return False


def _render_producto_form(request, *, template_full: str, modo: str, form, producto=None):
    ctx = {
        "form": form,
        "modo": modo,
        "producto": producto,
        "form_action_url": (
            reverse("producto_update", args=[producto.pk])
            if producto
            else reverse("producto_create")
        ),
        "modal_title": (
            f"Editar · {producto.codigo}" if producto else "Nuevo producto"
        ),
        "retorno_query": request.GET.urlencode(),
        **producto_listas_extra_context(
            producto,
            selected_ids=_producto_listas_seleccionadas_post(request),
        ),
    }
    if request.GET.get("modal") == "1":
        return render(request, "productos/form_fragment.html", ctx)
    return render(request, template_full, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def producto_create(request):
    if request.method == "POST":
        form = ProductoForm(request.POST)
        form_ok = form.is_valid()
        falta_lista = _producto_requiere_lista_precio(request, form)
        repetido = _producto_repite_en_misma_lista(request, form) if form_ok else False
        if form_ok and not falta_lista and not repetido:
            producto = form.save(commit=False)
            producto.precio_venta_editado = bool(form.cleaned_data.get("precio_venta_editado"))
            producto.save()
            sync_producto_listas_extras_from_post(request, producto)
            messages.success(request, f"Producto creado: {producto.codigo}")
            if request.GET.get("modal") == "1":
                return HttpResponse(
                    """
<div class="modal-body p-4">
  <div class="text-success fw-semibold mb-1">Producto guardado.</div>
  <div class="text-muted small">Ya podés cerrar esta ventana y seguir cargando la venta.</div>
</div>
<script>
  (function () {
    try {
      window.dispatchEvent(new CustomEvent("sirona:producto-guardado"));
    } catch (e) {}
    try {
      var el = document.getElementById("sironaModal");
      if (el && typeof bootstrap !== "undefined") {
        var m = bootstrap.Modal.getInstance(el);
        if (m) m.hide();
      }
    } catch (e) {}
  })();
</script>
""",
                    content_type="text/html",
                )
            return _redirect_productos_con_filtros(request)
    else:
        form = ProductoForm()
    return _render_producto_form(request, template_full="productos/form.html", modo="nuevo", form=form)


@login_required
@require_http_methods(["GET", "POST"])
def producto_update(request, pk: int):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        form = ProductoForm(request.POST, instance=producto)
        form_ok = form.is_valid()
        falta_lista = _producto_requiere_lista_precio(request, form)
        repetido = _producto_repite_en_misma_lista(request, form, producto) if form_ok else False
        if form_ok and not falta_lista and not repetido:
            producto = form.save(commit=False)
            producto.precio_venta_editado = bool(form.cleaned_data.get("precio_venta_editado"))
            producto.save()
            sync_producto_listas_extras_from_post(request, producto)
            messages.success(request, f"Producto actualizado: {producto.codigo}")
            if request.GET.get("modal") == "1":
                return HttpResponse(
                    """
<div class="modal-body p-4">
  <div class="text-success fw-semibold mb-1">Producto guardado.</div>
  <div class="text-muted small">Podés cerrar esta ventana.</div>
</div>
<script>
  (function () {
    try {
      window.dispatchEvent(new CustomEvent("sirona:producto-guardado"));
    } catch (e) {}
    try {
      var el = document.getElementById("sironaModal");
      if (el && typeof bootstrap !== "undefined") {
        var m = bootstrap.Modal.getInstance(el);
        if (m) m.hide();
      }
    } catch (e) {}
  })();
</script>
""",
                    content_type="text/html",
                )
            return _redirect_productos_con_filtros(request)
    else:
        form = ProductoForm(instance=producto)
    return _render_producto_form(
        request,
        template_full="productos/form.html",
        modo="editar",
        form=form,
        producto=producto,
    )


@login_required
@require_http_methods(["POST"])
def producto_inline_update(request, pk: int):
    """Guardado desde la tabla (edición en línea). No sincroniza listas de rubro extra."""
    producto = get_object_or_404(Producto, pk=pk)
    form = ProductoForm(request.POST, instance=producto)
    if form.is_valid():
        producto = form.save(commit=False)
        producto.precio_venta_editado = bool(form.cleaned_data.get("precio_venta_editado"))
        producto.save()
        messages.success(request, f"Producto actualizado: {producto.codigo}")
    else:
        parts = []
        for field, errs in form.errors.items():
            parts.append(f"{field}: {', '.join(str(e) for e in errs)}")
        messages.error(
            request,
            "No se pudo guardar desde la tabla. " + (" · ".join(parts) if parts else "Revisá los valores."),
        )
    return _redirect_productos_con_filtros(request)


@staff_required
@require_http_methods(["POST"])
def producto_delete(request, pk: int):
    producto = get_object_or_404(Producto, pk=pk)
    codigo = producto.codigo
    producto.delete()
    messages.success(request, f"Producto eliminado: {codigo}")
    return _redirect_productos_con_filtros(request)


@staff_required
@require_http_methods(["POST"])
def producto_toggle_habilitado(request, pk: int):
    producto = get_object_or_404(Producto, pk=pk)
    if not producto.habilitado and producto.stock <= 0:
        messages.warning(request, "No se puede habilitar un producto sin stock.")
        return _redirect_productos_con_filtros(request)
    producto.habilitado = not producto.habilitado
    if not producto.habilitado:
        producto.en_lista_precios = False
    producto.save()
    return _redirect_productos_con_filtros(request)


@staff_required
@require_http_methods(["POST"])
def producto_toggle_lista(request, pk: int):
    producto = get_object_or_404(Producto, pk=pk)
    if not producto.habilitado:
        messages.warning(request, "No podés poner en lista un producto deshabilitado.")
        return _redirect_productos_con_filtros(request)
    producto.en_lista_precios = request.POST.get("set_lista") == "1"
    producto.save(update_fields=["en_lista_precios"])
    return _redirect_productos_con_filtros(request)


@staff_required
@require_http_methods(["POST"])
def productos_acciones_masa(request):
    """Habilitar / deshabilitar / lista PDF sobre varios productos seleccionados."""
    accion = (request.POST.get("accion") or "").strip()
    ids = sorted({int(x) for x in request.POST.getlist("producto_id") if str(x).isdigit()})
    if not ids:
        messages.warning(request, "Seleccioná al menos un producto.")
        return _redirect_productos_con_filtros(request)

    existentes = set(Producto.objects.filter(pk__in=ids).values_list("pk", flat=True))
    ids = [i for i in ids if i in existentes]
    if not ids:
        messages.error(request, "No se encontraron productos válidos para la acción.")
        return _redirect_productos_con_filtros(request)

    if accion == "habilitar":
        sin_stock = Producto.objects.filter(pk__in=ids, habilitado=False, stock__lte=0).count()
        n = Producto.objects.filter(pk__in=ids, habilitado=False, stock__gt=0).update(habilitado=True)
        if n:
            messages.success(request, f"Se habilitaron {n} producto(s).")
        if sin_stock:
            messages.warning(
                request,
                f"No se habilitaron {sin_stock} producto(s) sin stock (o revisá el stock antes).",
            )
        if not n and not sin_stock:
            messages.info(request, "Los productos elegidos ya estaban habilitados.")
    elif accion == "deshabilitar":
        n = Producto.objects.filter(pk__in=ids).update(habilitado=False, en_lista_precios=False)
        messages.success(request, f"Se deshabilitaron {n} producto(s).")
    elif accion == "lista_precio":
        raw_lista = (request.POST.get("lista_id") or "").strip()
        modo = (request.POST.get("lista_modo") or "").strip()  # add/remove
        if not raw_lista.isdigit():
            messages.error(request, "Elegí una lista de precio.")
            return _redirect_productos_con_filtros(request)
        if modo not in ("add", "remove"):
            messages.error(request, "Acción de lista no válida.")
            return _redirect_productos_con_filtros(request)

        lista = ListaPrecios.objects.filter(pk=int(raw_lista)).first()
        if not lista:
            messages.error(request, "La lista de precio no existe.")
            return _redirect_productos_con_filtros(request)

        if lista.es_farmacia:
            # Farmacia: marca booleana en el producto.
            if modo == "add":
                deshab = Producto.objects.filter(pk__in=ids, habilitado=False).count()
                n = Producto.objects.filter(pk__in=ids, habilitado=True).update(en_lista_precios=True)
                messages.success(request, f"{n} producto(s) agregados a «{lista.nombre}».")
                if deshab:
                    messages.info(
                        request,
                        f"{deshab} producto(s) deshabilitados no se agregan a la lista hasta habilitarlos.",
                    )
            else:
                n = Producto.objects.filter(pk__in=ids).update(en_lista_precios=False)
                messages.success(request, f"Se quitó «{lista.nombre}» en {n} producto(s).")
            return _redirect_productos_con_filtros(request)

        # Listas de rubro: through table con precio.
        if modo == "remove":
            n, _ = ListaPrecioItem.objects.filter(lista_id=lista.pk, producto_id__in=ids).delete()
            messages.success(request, f"Se quitaron {n} producto(s) de «{lista.nombre}».")
            return _redirect_productos_con_filtros(request)

        # add: detectar conflictos (ya existe item)
        productos = list(Producto.objects.filter(pk__in=ids).order_by("descripcion", "codigo"))
        existentes = {
            it.producto_id: it
            for it in ListaPrecioItem.objects.filter(lista_id=lista.pk, producto_id__in=ids).select_related("producto")
        }
        conflictos = [p for p in productos if p.pk in existentes]

        if conflictos and request.POST.get("resolve_present") != "1":
            # Mostrar pantalla para decidir qué precio queda para los que ya estaban.
            filas = []
            for p in conflictos:
                it = existentes[p.pk]
                filas.append(
                    {
                        "producto": p,
                        "precio_existente": q2(it.precio_venta),
                        "precio_nuevo": q2(p.precio_venta),
                    }
                )
            rq = (request.POST.get("retorno_query") or "").strip()
            if not rq:
                rp = {
                    "q": request.POST.get("retorno_q") or "",
                    "tipo": request.POST.get("retorno_tipo") or "",
                    "proveedor": request.POST.get("retorno_proveedor") or "",
                    "lista": request.POST.get("retorno_lista") or "",
                    "estado": request.POST.get("retorno_estado") or "",
                    "ingreso": request.POST.get("retorno_ingreso") or "",
                }
                ponly = {k: v for k, v in rp.items() if v}
                pg = (request.POST.get("retorno_page") or "").strip()
                if pg.isdigit():
                    ponly["page"] = pg
                if ponly:
                    rq = urlencode(ponly)
            return render(
                request,
                "productos/acciones_masa_lista_confirmar.html",
                {
                    "lista": lista,
                    "ids": ids,
                    "conflictos": filas,
                    "total": len(ids),
                    "retorno_query": rq,
                },
            )

        # Aplicar: crear faltantes y resolver conflictos según elección.
        creados = 0
        mantenidos = 0
        sobrescritos = 0
        with transaction.atomic():
            for p in productos:
                it = existentes.get(p.pk)
                if it is None:
                    ListaPrecioItem.objects.create(
                        lista_id=lista.pk,
                        producto_id=p.pk,
                        precio_venta=q2(p.precio_venta),
                    )
                    creados += 1
                    continue
                choice = (request.POST.get(f"conf_{p.pk}") or "keep").strip()
                if choice == "overwrite":
                    ListaPrecioItem.objects.filter(pk=it.pk).update(precio_venta=q2(p.precio_venta))
                    sobrescritos += 1
                else:
                    mantenidos += 1

        if creados:
            messages.success(request, f"Se agregaron {creados} producto(s) a «{lista.nombre}».")
        if sobrescritos:
            messages.warning(request, f"Se actualizaron {sobrescritos} precio(s) en «{lista.nombre}».")
        if mantenidos:
            messages.info(request, f"{mantenidos} producto(s) ya estaban en «{lista.nombre}» (se mantuvo su precio).")
    elif accion == "eliminar":
        ok = 0
        protegidos = []
        for pid in ids:
            try:
                Producto.objects.get(pk=pid).delete()
                ok += 1
            except ProtectedError:
                protegidos.append(str(pid))
        if ok:
            messages.success(request, f"Se eliminaron {ok} producto(s).")
        if protegidos:
            messages.warning(
                request,
                "No se pudieron eliminar algunos ítems porque están referenciados en ventas, presupuestos u otros registros "
                f"(ids: {', '.join(protegidos)}).",
            )
        if not ok and not protegidos:
            messages.info(request, "No hubo productos para eliminar.")
    else:
        messages.error(request, "Acción no reconocida.")

    return _redirect_productos_con_filtros(request)


@login_required
@require_http_methods(["GET", "POST"])
def productos_import_excel(request):
    if request.method == "GET":
        return render(request, "productos/import_excel.html")

    f = request.FILES.get("archivo")
    if not f:
        return HttpResponseBadRequest("Falta archivo.")

    name = (getattr(f, "name", "") or "").lower()
    if not name.endswith(".xlsx"):
        messages.error(
            request,
            "El archivo debe ser Excel en formato .xlsx (Excel 2007 o posterior). "
            "Si tenés .xls, abrilo en Excel y guardalo como .xlsx.",
        )
        return redirect("productos_import_excel")

    max_mb = int(os.environ.get("SIRONA_MAX_XLSX_UPLOAD_MB", "25"))
    if getattr(f, "size", 0) and f.size > (max_mb * 1024 * 1024):
        messages.error(request, f"El archivo es demasiado grande (máx. {max_mb} MB).")
        return redirect("productos_import_excel")

    try:
        raw = f.read()
        if not raw:
            messages.error(request, "El archivo está vacío.")
            return redirect("productos_import_excel")
        # XLSX es un ZIP; validación rápida para evitar basura/errores raros.
        if not raw.startswith(b"PK"):
            messages.error(request, "El archivo no parece ser un .xlsx válido.")
            return redirect("productos_import_excel")
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as exc:
        detalle = f" Detalle: {exc}" if getattr(request.user, "is_staff", False) else ""
        messages.error(
            request,
            "No se pudo leer el archivo. Comprobá que sea un .xlsx válido." + detalle,
        )
        return redirect("productos_import_excel")

    ws = wb.active

    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    colmap: dict[str, int] | None = None
    data_start_row = 1
    if row1 and _es_fila_encabezado_productos(row1):
        colmap = _construir_mapa_columnas_import(row1)
        data_start_row = 2

    # Columnas: codigo(opcional), descripcion, tipo, costo, porcentaje_ganancia(opcional), precio_venta(opcional), stock(opcional), fecha_vencimiento(opcional)
    # Si la fila 1 tiene encabezados reconocidos, las columnas se ubican por título (orden libre).
    # Tipo: flexible — ver _resolver_tipo_producto
    #
    # Códigos que ya existen: no se pisan; se guardan en sesión y se muestran en un resumen para elegir.
    creados = 0
    conflictos: list[dict] = []

    try:
        with transaction.atomic():
            for i, row in enumerate(
                ws.iter_rows(min_row=data_start_row, values_only=True),
                start=data_start_row,
            ):
                if not row:
                    continue
                if all(v is None or _celda_texto(v) == "" for v in row):
                    continue

                codigo = _codigo_desde_celda(_celda_import(row, "codigo", colmap))
                descripcion = _celda_texto(_celda_import(row, "descripcion", colmap))
                tipo_raw = _celda_texto(_celda_import(row, "tipo", colmap))

                costo = _parse_decimal_celda(_celda_import(row, "costo", colmap), default=Decimal("0"))
                pct = _parse_decimal_celda(
                    _celda_import(row, "porcentaje_ganancia", colmap),
                    default=Decimal("30.00"),
                )
                precio = _parse_opcional_decimal(_celda_import(row, "precio_venta", colmap))

                stock = _parse_stock_importacion(_celda_import(row, "stock", colmap), i)

                fecha_vencimiento = None
                fv_cell = _celda_import(row, "fecha_vencimiento", colmap)
                if fv_cell is not None and _celda_texto(fv_cell) != "":
                    v = fv_cell
                    if hasattr(v, "date") and hasattr(v, "hour"):
                        fecha_vencimiento = v.date()
                    elif isinstance(v, date) and not hasattr(v, "hour"):
                        fecha_vencimiento = v
                    else:
                        s = _celda_texto(v)
                        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
                            try:
                                fecha_vencimiento = datetime.strptime(s, fmt).date()
                                break
                            except ValueError:
                                continue
                        if fecha_vencimiento is None:
                            raise ValueError(
                                f"Fila {i}: fecha de vencimiento no reconocida ({s!r}). "
                                "Usá dd/mm/aaaa o el formato de fecha del modelo."
                            )

                if not descripcion:
                    continue

                if not tipo_raw:
                    raise ValueError(
                        f"Fila {i}: la columna «tipo» está vacía (tercera columna: MED, AC u OT). "
                        "Revisá que la fila 1 tenga encabezados y los datos empiecen en la fila 2, "
                        "sin columnas desplazadas. Podés descargar el modelo desde esta pantalla."
                    )

                tipo = _resolver_tipo_producto(tipo_raw)
                if not tipo:
                    raise ValueError(
                        f"Fila {i}: tipo no reconocido ({tipo_raw!r}). "
                        "Ejemplos: medicamento(s), accesorio(s), otros; o MED, AC, OT."
                    )

                defaults = {
                    "descripcion": descripcion,
                    "tipo": tipo,
                    "costo": costo,
                    "stock": stock,
                    "fecha_vencimiento": fecha_vencimiento,
                    "porcentaje_ganancia": pct,
                    # Farmacia/PDF y rubros: se definen después en cada producto (no por importación).
                    "en_lista_precios": False,
                }

                if precio is None:
                    defaults["precio_venta_editado"] = False
                else:
                    defaults["precio_venta"] = precio
                    defaults["precio_venta_editado"] = True

                if codigo:
                    if Producto.objects.filter(codigo=codigo).exists():
                        existente = Producto.objects.get(codigo=codigo)
                        conflictos.append(
                            {
                                "fila": i,
                                "codigo": codigo,
                                "producto_id": existente.pk,
                                "excel": _excel_snapshot_for_session(defaults, fecha_vencimiento=fecha_vencimiento),
                            }
                        )
                        continue
                    Producto.objects.update_or_create(codigo=codigo, defaults=defaults)
                    creados += 1
                else:
                    obj = Producto(**defaults)
                    obj.save()
                    creados += 1
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("productos_import_excel")

    if conflictos:
        request.session[IMPORT_EXCEL_CONFLICTS_KEY] = {"items": conflictos}
        request.session.modified = True
        messages.success(
            request,
            f"Se cargaron {creados} producto(s) nuevo(s). "
            f"Hay {len(conflictos)} fila(s) con código ya existente: revisá el resumen y elegí qué datos conservar.",
        )
        return redirect("productos_import_excel_resumen")

    messages.success(request, f"Importación OK. Productos nuevos cargados: {creados}.")
    return redirect("productos_list")


@login_required
@require_http_methods(["GET", "POST"])
def productos_import_excel_resumen(request):
    """Tras importar, permite elegir por cada código duplicado si se aplica la fila del Excel o se mantiene la base."""
    payload = request.session.get(IMPORT_EXCEL_CONFLICTS_KEY) or {}
    items = list(payload.get("items") or [])

    if request.method == "POST":
        if not items:
            messages.info(request, "No había decisiones pendientes.")
            return redirect("productos_import_excel")
        actualizados = 0
        for it in items:
            pid = it.get("producto_id")
            if not pid:
                continue
            choice = (request.POST.get(f"resolver_{pid}") or "mantener").strip().lower()
            if choice == "excel":
                try:
                    p = Producto.objects.get(pk=pid)
                except Producto.DoesNotExist:
                    continue
                _aplicar_snapshot_excel_a_producto(p, it["excel"])
                actualizados += 1
        request.session.pop(IMPORT_EXCEL_CONFLICTS_KEY, None)
        messages.success(
            request,
            f"Listo. Se actualizaron {actualizados} producto(s) con los datos del Excel. "
            "El resto se mantuvo como estaba en la base.",
        )
        return redirect("productos_list")

    if not items:
        messages.info(request, "No hay un resumen de importación pendiente. Volvé a importar un archivo si hace falta.")
        return redirect("productos_import_excel")

    ids = [x["producto_id"] for x in items if x.get("producto_id")]
    productos = {p.pk: p for p in Producto.objects.filter(pk__in=ids)}
    filas = []
    for it in items:
        pid = it.get("producto_id")
        filas.append({**it, "producto": productos.get(pid) if pid else None})

    return render(
        request,
        "productos/import_excel_resumen.html",
        {"filas": filas},
    )


@login_required
@require_http_methods(["GET"])
def productos_import_excel_modelo(request):
    headers = [
        "codigo",
        "descripcion",
        "tipo",
        "costo",
        "porcentaje_ganancia",
        "precio_venta",
        "stock",
        "fecha_vencimiento",
    ]
    ejemplo = [
        [
            "",
            "Paracetamol 500mg",
            "MED",
            "100.00",
            "30.00",
            "",
            "0",
            "31/12/26",
        ]
    ]
    return xlsx_response("modelo_import_productos", [("Productos", headers, ejemplo)])


@login_required
def productos_export_pdf(request):
    incluir_stock = request.GET.get("stock") == "1"
    productos = list(
        Producto.objects.filter(en_lista_precios=True, habilitado=True).order_by(
            "tipo", "descripcion", "codigo"
        )
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = platypus_membrete("Lista de precios", doc.width, styles)

    headers = ["Código", "Tipo", "Descripción"]
    if incluir_stock:
        headers.append("Stock")
    headers.append("Precio")

    data = [headers]
    for p in productos:
        desc = p.descripcion
        if len(desc) > 85:
            desc = desc[:82] + "..."
        row = [p.codigo, p.get_tipo_display(), desc]
        if incluir_stock:
            row.append(str(p.stock))
        row.append(format_monto_ars(p.precio_venta))
        data.append(row)

    if len(data) == 1:
        row = ["—", "—", "—"]
        if incluir_stock:
            row.append("—")
        row.append("—")
        data.append(row)

    tw = doc.width
    if incluir_stock:
        col_w = [tw * 0.14, tw * 0.17, tw * 0.39, tw * 0.10, tw * 0.20]
    else:
        col_w = [tw * 0.15, tw * 0.18, tw * 0.45, tw * 0.22]

    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0097B2")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f9fb")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    fecha = timezone.localtime().strftime("%d-%m-%Y")
    filename = f"Lista_Precios_Sirona_{fecha}.pdf"
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type="application/pdf",
    )

