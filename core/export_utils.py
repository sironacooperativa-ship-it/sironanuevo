"""Exportación Excel (openpyxl) y PDF (reportlab) para listados."""
from __future__ import annotations

from io import BytesIO
from typing import Any
from xml.sax.saxutils import escape

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .pdf_membrete import platypus_membrete


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def xlsx_response(filename_base: str, sheets: list[tuple[str, list[str], list[list[Any]]]]) -> HttpResponse:
    wb = Workbook()
    first = True
    for sheet_name, headers, rows in sheets:
        title = (sheet_name or "Hoja1")[:31]
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([_cell_str(c) for c in row])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in filename_base)[:80]
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{safe}.xlsx"'
    return resp


def pdf_response(
    filename_base: str,
    doc_title: str,
    sections: list[tuple[str, list[str], list[list[Any]]]],
    *,
    body_fontsize: int = 7,
) -> HttpResponse:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    page_w = landscape(A4)[0] - 20 * mm
    story: list[Any] = []
    story.extend(platypus_membrete(doc_title, page_w, styles))

    for sec_title, headers, rows in sections:
        story.append(Paragraph(f"<b>{escape(sec_title)}</b>", styles["Heading3"]))
        data = [headers] + [[_cell_str(x) for x in row] for row in rows]
        if len(data) == 1:
            data.append([""] * len(headers))
        ncols = len(headers)
        col_w = page_w / max(ncols, 1)
        t = Table(data, colWidths=[col_w] * ncols, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0097B2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), body_fontsize),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f9fb")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 12))

    doc.build(story)
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in filename_base)[:80]
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{safe}.pdf"'
    return resp


def parse_export(request) -> str:
    if request.method != "GET":
        return ""
    return (request.GET.get("export") or "").lower()
